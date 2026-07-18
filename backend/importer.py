import io
import re
from typing import Any, Optional
import openpyxl
from database import Database

# Mapeamento de colunas do Excel para campos internos
COLUNAS_BATISMO = {
    "FONTE (Código de refª)": "fonte",
    "FLS": "fls",
    "ANO": "ano",
    "Nº ORDEM": "nr_ordem",
    "NOME": "nome",
    "DATA NASC": "data_nasc",
    "LOCAL NASCIMENTO": "local_nascimento",
    "PAI": "pai",
    "MÃE": "mae",
    "AVÔ PATERNO": "avo_paterno",
    "AVÓ PATERNA": "avo_paterna",
    "AVÔ MATERNO": "avo_materno",
    "AVÓ MATERNA": "avo_materna",
    "NOTAS": "notas",
}

COLUNAS_CASAMENTO = {
    "FONTE (Código de refª)": "fonte",
    "FLS": "fls",
    "ANO": "ano",
    "Nº ORDEM": "nr_ordem",
    "DATA": "data",
    "NOIVO": "noivo",
    "IDADE/DNASC NOIVO": "idade_dnasc_noivo",
    "NAT. NOIVO": "nat_noivo",
    "NOIVA": "noiva",
    "IDADE/DNASC NOIVA": "idade_dnasc_noiva",
    "NAT. NOIVA": "nat_noiva",
    "RESID": "residencia",
    "PAI NOIVO": "pai_noivo",
    "NAT PAI Nº": "nat_pai_noivo",
    "MÃE NOIVO": "mae_noivo",
    "NAT MÃE Nº": "nat_mae_noivo",
    "PAI NOIVA": "pai_noiva",
    "NAT PAI Nª": "nat_pai_noiva",
    "MÃE NOIVA": "mae_noiva",
    "NAT MÃE Nª": "nat_mae_noiva",
    "AVÔ PATERNO Nº": "avo_paterno_noivo",
    "AVÓ PATERNA Nº": "avo_paterna_noivo",
    "AVÔ MATERNO Nº": "avo_materno_noivo",
    "AVÓ MATERNA Nº": "avo_materna_noivo",
    "AVÔ PATERNO Nª": "avo_paterno_noiva",
    "AVÓ PATERNA Nª": "avo_paterna_noiva",
    "AVÔ MATERNO Nª": "avo_materno_noiva",
    "AVÓ MATERNA Nª": "avo_materna_noiva",
    "TESTEMUNHA 1": "testemunha1",
    "TESTEMUNHA 2": "testemunha2",
    "NOTAS": "notas",
}

COLUNAS_OBITO = {
    "FONTE (Código de refª)": "fonte",
    "FLS": "fls",
    "ANO": "ano",
    "Nº ORDEM": "nr_ordem",
    "NOME": "nome",
    "DATA ÓBITO": "data_obito",
    "LOCAL F.": "local_falecimento",
    "IDADE": "idade",
    "PAI": "pai",
    "NAT PAI": "nat_pai",
    "MÃE": "mae",
    "NAT MÃE": "nat_mae",
    "NOTAS": "notas",
}

MAPA_COLUNAS = {
    "batismo": COLUNAS_BATISMO,
    "casamento": COLUNAS_CASAMENTO,
    "obito": COLUNAS_OBITO,
}

CAMPOS_OBRIGATORIOS = {
    "batismo": ["nome", "ano"],
    "casamento": ["noivo", "noiva", "ano"],
    "obito": ["nome", "ano"],
}

ANO_MIN = 1500
ANO_MAX = 2100


def limpar(val: Any) -> str:
    if val is None:
        return "n/d"
    s = str(val).strip()
    return s if s else "n/d"


def validar_ano(val: Any) -> tuple[Optional[int], Optional[str]]:
    """Devolve (ano_int, aviso) onde aviso é None se OK."""
    if val is None:
        return None, "Ano em falta"
    try:
        ano = int(float(str(val)))
    except (ValueError, TypeError):
        return None, f"Ano inválido: '{val}'"
    if not (ANO_MIN <= ano <= ANO_MAX):
        return ano, f"Ano fora do intervalo esperado ({ANO_MIN}-{ANO_MAX}): {ano}"
    return ano, None


def _chave_registo(reg: dict, tipo: str) -> tuple:
    """
    Gera uma chave de identificação única para um registo.
    Prioridade 1: fonte + nr_ordem (referência de arquivo — univocamente identificadora).
    Prioridade 2: campos biográficos.
    """
    fonte    = (reg.get("fonte") or "").strip().lower()
    nr_ordem = (reg.get("nr_ordem") or "").strip().lower()

    if fonte and fonte != "n/d" and nr_ordem and nr_ordem != "n/d":
        return (tipo, "ref", fonte, nr_ordem)

    ano = str(reg.get("ano") or "")
    if tipo == "batismo":
        nome = (reg.get("nome") or "").strip().lower()
        pai  = (reg.get("pai")  or "").strip().lower()
        mae  = (reg.get("mae")  or "").strip().lower()
        return (tipo, "bio", ano, nome, pai, mae)
    elif tipo == "casamento":
        noivo = (reg.get("noivo") or "").strip().lower()
        noiva = (reg.get("noiva") or "").strip().lower()
        return (tipo, "bio", ano, noivo, noiva)
    else:
        nome = (reg.get("nome") or "").strip().lower()
        pai  = (reg.get("pai")  or "").strip().lower()
        mae  = (reg.get("mae")  or "").strip().lower()
        return (tipo, "bio", ano, nome, pai, mae)


def _descrever_registo(reg: dict, tipo: str) -> str:
    ano    = reg.get("ano") or "?"
    folha  = reg.get("_folha", "?")
    linha  = reg.get("_nr_linha", "?")
    loc    = f"[{folha}] linha {linha}"
    if tipo == "batismo":
        return f"{loc}: {reg.get('nome') or '?'} ({ano})"
    elif tipo == "casamento":
        return f"{loc}: {reg.get('noivo') or '?'} & {reg.get('noiva') or '?'} ({ano})"
    else:
        return f"{loc}: {reg.get('nome') or '?'} ({ano})"


class ExcelImporter:
    def __init__(self, db: Database):
        self.db = db

    def validar_e_importar(
        self, conteudo: bytes, tipo: str, nome_ficheiro: str, dry_run: bool,freguesia: str = None, modo_actualizacao: bool = False
    ) -> dict:
        mapa = MAPA_COLUNAS[tipo]
        obrigatorios = CAMPOS_OBRIGATORIOS[tipo]

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(conteudo), read_only=True, data_only=True
            )
        except Exception as e:
            return {"sucesso": False, "erro": f"Não foi possível abrir o ficheiro: {e}"}

        todos_registos = []
        todos_avisos = []
        erros_criticos = []

        for nome_folha in wb.sheetnames:
            ws = wb[nome_folha]
            registos_folha, avisos_folha, erros_folha = self._processar_folha(
                ws, nome_folha, tipo, mapa, obrigatorios
            )
            todos_registos.extend(registos_folha)
            todos_avisos.extend(avisos_folha)
            erros_criticos.extend(erros_folha)

        wb.close()

        # ── Detecção de duplicados ────────────────────────────────────────────

        # 1. Duplicados dentro do próprio ficheiro
        dup_intra, registos_unicos = self._detectar_duplicados_intra(todos_registos, tipo)
        for desc in dup_intra:
            todos_avisos.append(f"Duplicado no ficheiro: {desc}")

        # 2. Duplicados contra a base de dados existente
        dup_bd, registos_novos = self._detectar_duplicados_bd(registos_unicos, tipo)
        registos_unicos_antes_bd = registos_unicos  # guardar para actualização
        for desc in dup_bd:
            todos_avisos.append(f"Já existe na BD: {desc}")

        total_duplicados = len(dup_intra) + len(dup_bd)

        resumo = {
            "sucesso": True,
            "ficheiro": nome_ficheiro,
            "tipo": tipo,
            "total_registos": len(todos_registos),
            "total_novos": len(registos_novos),
            "total_duplicados": total_duplicados,
            "total_duplicados_intra": len(dup_intra),
            "total_duplicados_bd": len(dup_bd),
            "total_avisos": len(todos_avisos),
            "total_erros": len(erros_criticos),
            "avisos": todos_avisos[:100],  # limitar para não sobrecarregar a resposta
            "erros": erros_criticos[:50],
            "dry_run": dry_run,
        }

        if not dry_run and not erros_criticos:
            if modo_actualizacao and dup_bd:
                total_actualizados = self._actualizar_duplicados_bd(registos_unicos_antes_bd, tipo)
            else:
                total_actualizados = 0
                
            upload_id = self.db.criar_upload(
                nome_ficheiro, tipo, len(registos_novos), len(todos_avisos), freguesia
            )
            # Após inserir, extrair e guardar código do arquivo
            codigo = self.db.extrair_codigo_adist(upload_id, tipo)
            if codigo:
                self.db.actualizar_codigo_adist(upload_id, codigo)
            if tipo == "batismo":
                self.db.inserir_batismos(registos_novos, upload_id)
            elif tipo == "casamento":
                self.db.inserir_casamentos(registos_novos, upload_id)
            elif tipo == "obito":
                self.db.inserir_obitos(registos_novos, upload_id)
            resumo["upload_id"] = upload_id
            resumo["total_actualizados"] = total_actualizados
            
            partes = []
            if registos_novos:
                partes.append(f"{len(registos_novos)} registos novos importados")
            if total_actualizados:
                partes.append(f"{total_actualizados} atualizados")
            if total_duplicados and not modo_actualizacao:
                partes.append(f"{total_duplicados} duplicado(s) ignorados")
            resumo["mensagem"] = ". ".join(partes) + "." if partes else "Nenhuma alteração efetuada."
        elif not dry_run and erros_criticos:
            resumo["sucesso"] = False
            resumo["mensagem"] = "Importação cancelada devido a erros críticos."

        return resumo

    def _actualizar_duplicados_bd(self, registos: list, tipo: str) -> int:
        total = 0
        for reg in registos:
            chave = _chave_registo(reg, tipo)
            if chave[1] == "ref":
                fonte    = reg.get("fonte", "")
                nr_ordem = reg.get("nr_ordem", "")
                if self.db.actualizar_registo_por_ref(tipo, fonte, nr_ordem, reg):
                    total += 1
            else:
                if self.db.actualizar_registo_por_bio(tipo, chave, reg):
                    total += 1
        return total

    def _detectar_duplicados_intra(
        self, registos: list, tipo: str
    ) -> tuple[list[str], list[dict]]:
        """Detecta duplicados dentro do ficheiro (incluindo entre folhas diferentes)."""
        visto = {}
        unicos = []
        duplicados = []
        for reg in registos:
            chave = _chave_registo(reg, tipo)
            if chave in visto:
                duplicados.append(_descrever_registo(reg, tipo))
            else:
                visto[chave] = True
                unicos.append(reg)
        return duplicados, unicos

    def _detectar_duplicados_bd(
        self, registos: list, tipo: str
    ) -> tuple[list[str], list[dict]]:
        """Verifica quais registos já existem na base de dados."""
        if not registos:
            return [], []

        refs = [(i, r) for i, r in enumerate(registos)
                if _chave_registo(r, tipo)[1] == "ref"]
        bios = [(i, r) for i, r in enumerate(registos)
                if _chave_registo(r, tipo)[1] == "bio"]

        existentes_idx = set()
        duplicados = []

        if refs:
            pares = [(r.get("fonte", ""), r.get("nr_ordem", "")) for _, r in refs]
            resultados = self.db.verificar_existencia_por_ref(tipo, pares)
            for i, (idx, reg) in enumerate(refs):
                if resultados[i]:
                    existentes_idx.add(idx)
                    duplicados.append(_descrever_registo(reg, tipo))

        if bios:
            chaves = [_chave_registo(r, tipo) for _, r in bios]
            resultados = self.db.verificar_existencia_por_bio(tipo, chaves)
            for i, (idx, reg) in enumerate(bios):
                if resultados[i]:
                    existentes_idx.add(idx)
                    duplicados.append(_descrever_registo(reg, tipo))

        novos = [r for i, r in enumerate(registos) if i not in existentes_idx]
        return duplicados, novos

    def _processar_folha(self, ws, nome_folha, tipo, mapa, obrigatorios):
        registos = []
        avisos = []
        erros = []

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return registos, avisos, erros

        # Encontrar linha de cabeçalho (primeira linha não vazia)
        cabecalho_idx = None
        cabecalho = []
        for i, row in enumerate(rows):
            valores = [str(v).strip() if v is not None else "" for v in row]
            nao_vazios = [v for v in valores if v]
            if len(nao_vazios) >= 3:
                cabecalho_idx = i
                cabecalho = valores
                break

        if cabecalho_idx is None:
            avisos.append(f"[{nome_folha}] Folha sem cabeçalho reconhecível — ignorada.")
            return registos, avisos, erros

        # Mapear colunas do Excel para campos internos
        indice_campo = {}
        colunas_nao_reconhecidas = []
        for i, col in enumerate(cabecalho):
            col_norm = col.strip()
            if col_norm in mapa:
                indice_campo[mapa[col_norm]] = i
            elif col_norm:
                colunas_nao_reconhecidas.append(col_norm)

        if colunas_nao_reconhecidas:
            avisos.append(
                f"[{nome_folha}] Colunas não reconhecidas (serão ignoradas): "
                + ", ".join(colunas_nao_reconhecidas)
            )
        
        # Verificar colunas obrigatórias
        for campo in obrigatorios:
            if campo not in indice_campo:
                erros.append(f"[{nome_folha}] Coluna obrigatória em falta: '{campo}'")

        if erros:
            return registos, avisos, erros
        
        # Processar linhas de dados
        for nr_linha, row in enumerate(rows[cabecalho_idx + 1:], start=cabecalho_idx + 2):
            if all(v is None or str(v).strip() == "" for v in row):
                continue

            reg = {"_folha": nome_folha, "_nr_linha": nr_linha}
            for campo, idx in indice_campo.items():
                val = row[idx] if idx < len(row) else None
                if campo == "ano":
                    ano, aviso = validar_ano(val)
                    reg["ano"] = ano
                    if aviso:
                        avisos.append(f"[{nome_folha}] Linha {nr_linha}: {aviso}")
                else:
                    reg[campo] = limpar(val)

            # Verificar campos obrigatórios
            for campo in obrigatorios:
                if not reg.get(campo):
                    avisos.append(
                        f"[{nome_folha}] Linha {nr_linha}: campo '{campo}' em falta ou vazio"
                    )

            registos.append(reg)

        return registos, avisos, erros
